#!/usr/bin/env python3
"""
CCCF — 中国计算机学会通讯 下载 & 跟踪工具
========================================
自动下载《中国计算机学会通讯》(CCCF) 期刊 PDF，
支持《计算》主刊 + 《快讯》简报版，生成下载跟踪表。

用法:
  python3 cccf_download.py --year 2025 --issues 1-12    # 下载全年
  python3 cccf_download.py --update                      # 智能更新缺期
  python3 cccf_download.py --list                        # 查看跟踪表
  python3 cccf_download.py --login                       # 更新 Cookie

跨平台: Windows / Linux / macOS
依赖: 需要已登录 CCF 数字图书馆的浏览器 Cookie
"""

import os
import sys
import json
import time
import random
import string
import argparse
from pathlib import Path

# ==================== 配置 ====================
CCCF_BASE_URL = "https://dl.ccf.org.cn"
CCCF_INDEX_URL = f"{CCCF_BASE_URL}/article/articleIndex.html?typeClassVal=cccf"
DOWNLOAD_URL = f"{CCCF_BASE_URL}/downloadPDF"

# 本地路径（跨平台兼容）
HOME_DIR = str(Path.home())
OUTPUT_DIR = os.path.join(HOME_DIR, "CCCF_Downloads")
CONFIG_DIR = os.path.join(HOME_DIR, ".cccf-downloader")
COOKIE_FILE = os.path.join(CONFIG_DIR, "cookies.json")
TRACK_FILE = os.path.join(OUTPUT_DIR, ".cccf_tracker.json")

# 已知的 issue ID 映射 (可用于无浏览器时的快速下载)
# 格式: {year_issue: {id, func}}
ISSUE_IDS = {
    "2025_12": {"id": "7819224295262208", "func": "downloadPdf"},
    "2025_11": {"id": "7778133632944128", "func": "downloadPdf"},
    "2025_10": {"id": "7738916185737216", "func": "downloadPdf"},
    "2025_9":  {"id": "7688081520199680", "func": "downloadPdf"},
    "2025_8":  {"id": "7651309677512704", "func": "downloadPdf"},
    "2025_7":  {"id": "7607710551230464", "func": "downloadPdf"},
    "2025_6":  {"id": "7561885813131264", "func": "downloadPdf"},
    "2025_5":  {"id": "7517739741415424", "func": "downloadPdf"},
    "2025_4":  {"id": "7468103024052224", "func": "downloadPdf"},
    "2025_3":  {"id": "7426723023472640", "func": "downloadPdf"},
    "2025_2":  {"id": "7387433688844288", "func": "downloadPdf"},
    "2025_1":  {"id": "7341080779884544", "func": "downloadPdf"},
    "2026_6":  {"id": "8080874787768320", "func": "downloadPdfByCccfType"},
    "2026_5":  {"id": "8036310061287424", "func": "downloadPdfByCccfType"},
    "2026_4":  {"id": "7993320593524736", "func": "downloadPdfByCccfType"},
    "2026_3":  {"id": "7952846815807488", "func": "downloadPdfByCccfType"},
    "2026_2":  {"id": "7897791162796032", "func": "downloadPdfByCccfType"},
    "2026_1":  {"id": "7863609767364608", "func": "downloadPdfByCccfType"},
}


# ==================== 工具函数 ====================







# ==================== 下载跟踪表 ====================

def load_tracker():
    """加载下载跟踪 JSON"""
    if not os.path.exists(TRACK_FILE):
        return {"years": {}}
    with open(TRACK_FILE) as f:
        return json.load(f)


def save_tracker(data):
    """保存下载跟踪 JSON"""
    os.makedirs(os.path.dirname(TRACK_FILE), exist_ok=True)
    with open(TRACK_FILE, "w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def update_tracker(results):
    """下载完成后更新跟踪表"""
    tracker = load_tracker()
    now = time.strftime("%Y-%m-%d %H:%M")

    for item in results:
        y = str(item["year"])
        i = str(item["issue"]).zfill(2)
        is_news = item.get("is_news", False)
        size = os.path.getsize(item["filepath"]) if os.path.exists(item["filepath"]) else 0

        if y not in tracker["years"]:
            tracker["years"][y] = {}

        if i not in tracker["years"][y]:
            tracker["years"][y][i] = {"计算": None, "快讯": None}

        type_key = "快讯" if is_news else "计算"
        tracker["years"][y][i][type_key] = {
            "downloaded_at": now,
            "size_bytes": size,
            "filename": os.path.basename(item["filepath"])
        }

    save_tracker(tracker)
    export_tracker_xlsx()
    return tracker


def get_available_issues():
    """获取 CCF 每年的最大期数（月刊，当前年=当前月，往年=12）"""
    import datetime
    this_year = datetime.datetime.now().year
    this_month = datetime.datetime.now().month
    result = {}
    for y in range(2005, this_year):
        result[str(y)] = 12  # 往年全年
    result[str(this_year)] = this_month  # 当年只到当前月
    # 脚本已知的最新年份数据（2025-2026 可确认）
    return result


def export_tracker_xlsx():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return
    tracker = load_tracker()
    available = get_available_issues()
    xlsx_path = os.path.join(OUTPUT_DIR, "CCCF_下载跟踪表.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CCCF下载跟踪表"
    
    DEEP_BLUE, HEADER_BG = "1B3A5C", "1F4E79"
    GREEN_LIGHT, YELLOW_LIGHT = "C6EFCE", "FFF2CC"
    GRAY_LIGHT, GRAY_MID, WHITE = "E8E8E8", "BFBFBF", "FFFFFF"
    
    T_BG_CALC = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    DONE_F = PatternFill(start_color=GREEN_LIGHT, end_color=GREEN_LIGHT, fill_type="solid")
    MISS_F = PatternFill(start_color=YELLOW_LIGHT, end_color=YELLOW_LIGHT, fill_type="solid")
    FUT_F = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
    HDR_F = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    
    THIN = Border(left=Side("thin","D0D0D0"),right=Side("thin","D0D0D0"),top=Side("thin","D0D0D0"),bottom=Side("thin","D0D0D0"))
    CC = Alignment(horizontal="center",vertical="center")
    
    # Title
    ws.merge_cells("A1:O1")
    c=ws.cell(row=1, column=1, value="CCCF 中国计算机学会通讯 - 下载跟踪表")
    c.font=Font(bold=True,size=14,color=DEEP_BLUE,name="微软雅黑")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=36
    
    # Header
    for ci,h in enumerate(["年份","类型","1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月","下载状态"],1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=Font(bold=True,color="FFFFFF",size=10,name="微软雅黑")
        c.fill=HDR_F;c.alignment=CC;c.border=THIN
    ws.row_dimensions[2].height=28
    
    years=sorted(tracker["years"].keys(),reverse=True)
    r=3
    for y in years:
        mx=available.get(y,12)
        co=no=0
        for i in range(1,mx+1):
            d=tracker["years"][y].get(str(i).zfill(2),{})
            if d.get("计算"):co+=1
            if d.get("快讯"):no+=1
        
        ws.cell(row=r,column=1,value=int(y)).font=Font(bold=True,size=12,color=DEEP_BLUE,name="微软雅黑")
        ws.cell(row=r,column=1).alignment=CC;ws.cell(row=r,column=1).border=THIN
        ct=ws.cell(row=r,column=2,value="📖 计算")
        ct.font=Font(bold=True,size=10,color="1F4E79",name="微软雅黑")
        ct.fill=T_BG_CALC;ct.alignment=CC;ct.border=THIN
        
        for i in range(1,13):
            c=ws.cell(row=r,column=i+2);c.alignment=CC;c.border=THIN
            d=tracker["years"][y].get(str(i).zfill(2),{})
            if i>mx:
                c.value="---";c.fill=FUT_F
            elif d.get("计算"):
                c.value="\u2705";c.fill=DONE_F
            else:
                c.value="\u2b1c";c.fill=MISS_F
        
        sc=ws.cell(row=r,column=15);sc.border=THIN;sc.alignment=CC
        if co==0:sc.value="📭 缺期";sc.font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑");sc.fill=MISS_F
        elif co<mx:sc.value=f"⏳ 部分 {co}/{mx}";sc.font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑");sc.fill=MISS_F
        else:sc.value=f"✅ 完整 {co}/{mx}";sc.font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑");sc.fill=DONE_F
        ws.row_dimensions[r].height=28
        
        r+=1
        ws.cell(row=r,column=1,value=int(y)).font=Font(bold=True,size=12,color=DEEP_BLUE,name="微软雅黑")
        ws.cell(row=r,column=1).alignment=CC;ws.cell(row=r,column=1).border=THIN
        ct=ws.cell(row=r,column=2,value="📬 快讯")
        ct.font=Font(size=10,color="5B9BD5",name="微软雅黑")
        ct.alignment=CC;ct.border=THIN
        
        for i in range(1,13):
            c=ws.cell(row=r,column=i+2);c.alignment=CC;c.border=THIN
            d=tracker["years"][y].get(str(i).zfill(2),{})
            if i>mx:c.value="---";c.fill=FUT_F
            elif d.get("快讯"):c.value="\u2705";c.fill=DONE_F
            else:c.value="\u2b1c"
        
        sc=ws.cell(row=r,column=15);sc.border=THIN;sc.alignment=CC
        if no==0:sc.value="📭 无快讯";sc.font=Font(size=9,color=GRAY_MID,name="微软雅黑")
        elif no<mx:sc.value=f"⏳ {no}/{mx}";sc.font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑");sc.fill=MISS_F
        else:sc.value=f"✅ {no}/{mx}";sc.font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑");sc.fill=DONE_F
        ws.row_dimensions[r].height=28
        r+=1
    
    r+=1
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=15)
    ws.cell(row=r,column=1,value=f"📊 统计：已下载 {sum(len(v) for v in tracker['years'].values())} 期 | 文件 {sum(1 for _ in __import__('pathlib').Path(OUTPUT_DIR).glob('CCCF_*.pdf')) if __import__('os').path.exists(OUTPUT_DIR) else 0} 个").font=Font(size=10,color=DEEP_BLUE,name="微软雅黑")
    r+=1
    
    last=None
    for y in sorted(tracker["years"].keys(),reverse=True):
        for i in sorted(tracker["years"][y].keys(),reverse=True):
            for t in ["计算","快讯"]:
                d=tracker["years"][y][i].get(t,{})
                if d and d.get("downloaded_at"):
                    if not last or d["downloaded_at"]>last:last=d["downloaded_at"]
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=15)
    ws.cell(row=r,column=1,value=f"🕐 最近下载：{last or '无'}").font=Font(size=10,color=DEEP_BLUE,name="微软雅黑")
    r+=2
    
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=15)
    ws.cell(row=r,column=1,value="📌 图例").font=Font(bold=True,size=10,color=DEEP_BLUE,name="微软雅黑")
    r+=1
    for e,d in [("\u2705 绿色","已下载"),("\u2b1c 黄色","缺期"),("--- 灰色","未出版"),("📖 计算","主刊"),("📬 快讯","简报版")]:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
        ws.cell(row=r,column=1,value=e).font=Font(size=9,color="666666",name="微软雅黑")
        ws.cell(row=r,column=1).alignment=Alignment(horizontal="right",vertical="center")
        ws.merge_cells(start_row=r,start_column=3,end_row=r,end_column=15)
        ws.cell(row=r,column=3,value=d).font=Font(size=9,color="666666",name="微软雅黑")
        r+=1
    
    ws.column_dimensions["A"].width=8
    ws.column_dimensions["B"].width=10
    for ci in range(3,15):ws.column_dimensions[get_column_letter(ci)].width=6.5
    ws.column_dimensions["O"].width=22
    ws.freeze_panes="A3"
    ws.sheet_properties.pageSetUpPr=openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth=1
    ws.page_setup.orientation="landscape"
    wb.save(xlsx_path)
def show_tracker_table():
    """以表格形式展示下载状态（每年两行：计算 + 快讯）"""
    tracker = load_tracker()
    available = get_available_issues()

    if not tracker["years"]:
        print("\n📭 暂无下载记录")
        print("   运行 --year 2025 --issues 1-12 开始下载")
        return

    header = "| 年份 | 类型 | 01 | 02 | 03 | 04 | 05 | 06 | 07 | 08 | 09 | 10 | 11 | 12 | 状态 |"
    sep =    "| :--- | :--- | :- | :- | :- | :- | :- | :- | :- | :- | :- | :- | :- | :- | :-- |"
    print(f"\n📊 **CCCF 下载跟踪表**\n")
    print(header)
    print(sep)

    years = sorted(tracker["years"].keys(), reverse=True)
    for y in years:
        max_issue = available.get(y, 12)

        # 计算行
        calc_cells = [f"| {y} ", "| 《计算》 "]
        news_cells = [f"| {y} ", "| 《快讯》 "]
        calc_ok = news_ok = 0

        for i in range(1, 13):
            key = str(i).zfill(2)
            idata = tracker["years"][y].get(key, {})

            if i > max_issue:
                calc_cells.append("| — ")
                news_cells.append("| — ")
            elif idata.get("计算"):
                calc_cells.append("| ● ")
                calc_ok += 1
            else:
                calc_cells.append("| ○ ")

            if i > max_issue:
                continue
            elif idata.get("快讯"):
                news_cells.append("| ● ")
                news_ok += 1
            else:
                news_cells.append("| ○ ")

        calc_total = min(max_issue, 12)
        if calc_ok == calc_total:
            calc_cells.append("| 完整 |")
        elif calc_ok == 0:
            calc_cells.append("| 缺期 |")
        else:
            calc_cells.append(f"| {calc_ok}/{calc_total} |")

        if news_ok == 0:
            news_cells.append("| 无 |")
        elif news_ok == min(max_issue, 12):
            news_cells.append("| 完整 |")
        else:
            news_cells.append(f"| {news_ok}/{calc_total} |")

        print("".join(calc_cells))
        print("".join(news_cells))

    total_issues = sum(len(v) for v in tracker["years"].values())
    print(f"\n📈 共计 **{total_issues} 期** | 图例: ●=已下载  ○=缺期  —=未出版")

    last = None
    for y in sorted(tracker["years"].keys(), reverse=True):
        for i in sorted(tracker["years"][y].keys(), reverse=True):
            for t in ["计算", "快讯"]:
                d = tracker["years"][y][i].get(t, {})
                if d and d.get("downloaded_at"):
                    if not last or d["downloaded_at"] > last:
                        last = d["downloaded_at"]
    if last:
        print(f"🕐 最近下载: {last}")


def list_downloaded():
    """列出已下载的期刊，以表格形式展示"""
    # 如果未使用 tracker 则用文件系统扫描回退
    tracker = load_tracker()
    if tracker["years"]:
        show_tracker_table()
        return

    # 回退：扫描文件系统
    if not os.path.exists(OUTPUT_DIR):
        print("暂无下载记录")
        return
    files = sorted(Path(OUTPUT_DIR).glob("CCCF_*.pdf"))
    if not files:
        print("暂无下载记录")
        return

    total_size = sum(f.stat().st_size for f in files)
    print(f"\n📚 已下载的 CCCF 期刊 ({len(files)} 期, {total_size//1024//1024}MB)")
    print("-" * 50)
    for f in files:
        name = f.stem
        parts = name.split("_")
        size = f.stat().st_size // 1024 // 1024
        print(f"  📄 {parts[1]}年 第{int(parts[2])}期  ({size}MB)")

def save_cookies_from_browser():
    """引导用户在浏览器中登录 CCF，保存 Cookie"""
    print("\n⚠️  需要登录 CCF 数字图书馆。请在浏览器中:")
    print("   1. 访问 https://dl.ccf.org.cn 并登录")
    print("   2. 登录成功后，复制下面命令的输出:")
    print()
    print("   在已登录的浏览器开发者工具(Console)中执行:")
    print()
    print('   copy(document.cookie)')
    print()
    print("   然后粘贴到这里:")
    cookie_str = input("   Cookie > ").strip()

    if not cookie_str:
        print("❌ 未提供 Cookie")
        return False

    # 解析 cookie string 为 JSON
    cookies = {}
    for item in cookie_str.split(";"):
        item = item.strip()
        if "=" in item:
            k, v = item.split("=", 1)
            cookies[k.strip()] = v.strip()

    with open(COOKIE_FILE, "w") as f:
        json.dump(cookies, f, indent=2)
    print(f"✅ Cookie 已保存到 {COOKIE_FILE}")
    return True


def load_cookies():
    """加载保存的 Cookie"""
    if not os.path.exists(COOKIE_FILE):
        return None
    with open(COOKIE_FILE) as f:
        return json.load(f)


def cookie_to_header(cookies):
    """将 cookies dict 转为 HTTP Header 字符串"""
    return "; ".join(f"{k}={v}" for k, v in cookies.items())


def get_issue_ids_from_web(year, issue, cookies):
    """通过网页获取指定期号的下载 ID（计算+快讯）"""
    import urllib.request
    import re

    headers = {
        "Cookie": cookie_to_header(cookies),
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
    }

    url = f"{CCCF_INDEX_URL}&year={year}&issue={issue}"
    req = urllib.request.Request(url, headers=headers)

    try:
        resp = urllib.request.urlopen(req)
        html = resp.read().decode("utf-8")

        # 检查登录状态
        if "登录" in html and "安全退出" not in html:
            print("⚠️  登录已过期，请重新登录")
            return None, None

        # 提取《计算》下载 ID（cccfType=01）
        calc_id = None
        m = re.search(r"downloadPdfByCccfType\('(\d+)\|01'", html)
        if m:
            calc_id = m.group(1)
        else:
            m = re.search(r"downloadPdf\((\d+),", html)
            if m:
                calc_id = m.group(1)

        # 提取《快讯》下载 ID（cccfType=02）
        news_id = None
        m = re.search(r"downloadPdfByCccfType\('(\d+)\|02'", html)
        if m:
            news_id = m.group(1)

        if not calc_id:
            print(f"⚠️  未找到 {year}年第{issue}期的下载 ID")
            return None, None

        return calc_id, news_id

    except Exception as e:
        print(f"❌ 网络请求失败: {e}")
        return None, None


def download_issue_pdf(issue_id, year, issue, cookies, output_dir, type_label="计算", cccf_type=None):
    """下载单期 PDF（支持《计算》和《快讯》）"""
    import urllib.request

    suffix = "" if type_label == "计算" else f"_{type_label}"
    filename = f"CCCF_{year}_{issue:02d}{suffix}.pdf"
    filepath = os.path.join(output_dir, filename)

    if os.path.exists(filepath) and os.path.getsize(filepath) > 1000000:
        print(f"  ⏭️  [{type_label}] 已存在 ({os.path.getsize(filepath)//1024//1024}MB)")
        return filepath

    url = f"{DOWNLOAD_URL}?id={issue_id}"
    if cccf_type:
        url += f"&cccfType={cccf_type}"
    headers = {
        "Cookie": cookie_to_header(cookies),
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"
    }

    req = urllib.request.Request(url, headers=headers)
    try:
        resp = urllib.request.urlopen(req)
        size = int(resp.headers.get("Content-Length", 0))
        print(f"  📥 [{type_label}] 下载中... ({size//1024//1024}MB)")

        with open(filepath, "wb") as f:
            f.write(resp.read())

        actual = os.path.getsize(filepath)
        print(f"  ✅ [{type_label}] 完成 ({actual//1024//1024}MB)")
        return filepath
    except Exception as e:
        print(f"  ❌ [{type_label}] 下载失败: {e}")
        return None


def download_issues(year_start, issue_start, issue_end, cookies):
    """下载指定年/期范围的 PDF"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    results = []
    for issue in range(issue_start, issue_end + 1):
        key = f"{year_start}_{issue}"

        # 获取下载 ID（计算+快讯）
        calc_id = news_id = None
        if key in ISSUE_IDS:
            calc_id = ISSUE_IDS[key]["id"]
            print(f"\n📄 CCCF {year_start}年第{issue}期 (已知ID)")
        else:
            print(f"\n📄 CCCF {year_start}年第{issue}期 (从网页获取ID)...")
            calc_id, news_id = get_issue_ids_from_web(year_start, issue, cookies)

        if not calc_id:
            print(f"  ⏭️  跳过")
            continue

        # 下载《计算》
        filepath = download_issue_pdf(calc_id, year_start, issue, cookies, OUTPUT_DIR, type_label="计算", cccf_type="01" if key not in ISSUE_IDS else None)
        if filepath:
            results.append({
                "year": year_start,
                "issue": issue,
                "filepath": filepath,
                "filename": f"CCCF_{year_start}_{issue:02d}.pdf"
            })

        # 下载《快讯》（如有）
        if news_id:
            print(f"  📋 发现《快讯》版本")
            news_path = download_issue_pdf(news_id, year_start, issue, cookies, OUTPUT_DIR, type_label="快讯", cccf_type="02")
            if news_path:
                results.append({
                    "year": year_start,
                    "issue": issue,
                    "filepath": news_path,
                    "filename": f"CCCF_{year_start}_{issue:02d}_快讯.pdf",
                    "is_news": True
                })
        else:
            print(f"  📋 本期无《快讯》版本")

    # 更新跟踪表
    update_tracker(results)

    return results


# ==================== Zotero 导入相关 ====================



def do_update():
    """智能检测最新可下载的期号，自动补全缺期"""
    import datetime

    cookies = load_cookies()
    if not cookies:
        print("⚠️  未找到已保存的 CCF 登录信息")
        return

    now = datetime.datetime.now()
    this_year = now.year
    this_month = now.month

    tracker = load_tracker()

    # 计算需要检查的年份范围：从已有记录的最早年到今年
    existing_years = [int(y) for y in tracker.get("years", {}).keys()]
    start_year = min(existing_years) if existing_years else this_year - 1

    print(f"\n{'='*50}")
    print(f"🔄 智能更新检查: 当前 {this_year}年{this_month}月")
    print(f"{'='*50}")

    all_downloaded = 0
    for year in range(start_year, this_year + 1):
        # 该年最多到哪一期
        if year == this_year:
            max_issue = this_month  # 今年只出到当前月
        else:
            max_issue = 12  # 往年全年

        missing = []
        for issue in range(1, max_issue + 1):
            key = str(issue).zfill(2)
            year_data = tracker.get("years", {}).get(str(year), {})
            issue_data = year_data.get(key, {})
            if not issue_data.get("计算"):
                missing.append(issue)

        if missing:
            print(f"\n📋 {year}年 缺期: 第{'、'.join(str(i) for i in missing)}期")
            results = download_issues(year, min(missing), max(missing), cookies)
            all_downloaded += len(results)
        else:
            print(f"  ✅ {year}年 已完整 (第1-{max_issue}期)")

    if all_downloaded == 0:
        print("\n✨ 所有已知期刊已是最新，无需下载")
    else:
        print(f"\n📥 本次更新: {all_downloaded} 个文件")


def main():
    parser = argparse.ArgumentParser(description="CCCF 下载 CCCF 下载 & Zotero 导入工具 跟踪工具")
    parser.add_argument("--year", type=int, help="年份 (如 2025)")
    parser.add_argument("--issues", type=str, help="期号范围 (如 1-6 或 1,3,5)")
    parser.add_argument("--download-only", action="store_true", help="仅下载 PDF，跳过 Zotero 导入")
    parser.add_argument("--list", "-l", action="store_true", help="查看下载状态跟踪表")
    parser.add_argument("--login", action="store_true", help="更新 CCF 登录 Cookie")
    parser.add_argument("--update", "-u", action="store_true", help="检查并下载最新缺期期刊")

    args = parser.parse_args()

    if args.list:
        show_tracker_table()
        return

    if args.login:
        save_cookies_from_browser()
        return

    if args.update:
        do_update()
        return

    if not args.year or not args.issues:
        parser.print_help()
        print("\n💡 常用示例:")
        print("  python3 cccf_download.py --year 2025 --issues 1-12")
        print("  python3 cccf_download.py --year 2026 --issues 1-6")
        print("  python3 cccf_download.py --list")
        print("  python3 cccf_download.py --login")
        return

    # 解析期号范围
    if "-" in args.issues:
        parts = args.issues.split("-")
        issue_start, issue_end = int(parts[0]), int(parts[1])
    else:
        issue_start = issue_end = int(args.issues)

    # 1. 加载/获取 Cookie
    cookies = load_cookies()
    if not cookies:
        print("⚠️  未找到已保存的 CCF 登录信息")
        if not save_cookies_from_browser():
            return
        cookies = load_cookies()

    # 2. 下载 PDF
    print(f"\n{'='*50}")
    print(f"📥 开始下载 CCCF {args.year}年 第{issue_start}-{issue_end}期")
    print(f"{'='*50}")

    results = download_issues(args.year, issue_start, issue_end, cookies)

    if not results:
        print("\n❌ 下载失败，可能是 Cookie 已过期")
        print("请运行: python3 cccf_download.py --login")
        return

    print(f"\n✅ 下载完成: {len(results)} 期")
    print(f"   保存到: {OUTPUT_DIR}")

    # 3. 导入 Zotero
    if not args.download_only:
        print(f"\n{'='*50}")
        print("📚 导入 Zotero")
        print(f"{'='*50}")
        import_to_zotero(results)

    print(f"\n{'='*50}")
    print("✅ 全部完成!")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
